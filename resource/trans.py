import openpyxl
import csv

# openpyxl.Workbook.encoding="cp936"
wb = openpyxl.load_workbook('USNews2021.xlsx')

sheet = wb[wb.sheetnames[0]]


csv_file = open('USNews2021.csv','w', newline='',encoding='utf-8')
writer = csv.writer(csv_file)

a=[]
for i in sheet.iter_rows(min_row=1, max_row=110, min_col=1,max_col=2):
    l = []
    for j in i:
        l.append(j.value)
    a.append(l)
    

dicc = {}
ind = open('QS2024.csv',encoding='utf-8')
reader = csv.reader(ind)
for row in reader :
    dicc[row[1]] = row[2]
ind = open('THE2024.csv',encoding='utf-8')
reader = csv.reader(ind)
for row in reader :
    dicc[row[1]] = row[2]
ind = open('ARWU2023.csv',encoding='utf-8')
reader = csv.reader(ind)
for row in reader :
    dicc[row[1]] = row[2]
ind = open('USNews2023.csv',encoding='utf-8')
reader = csv.reader(ind)
for row in reader :
    dicc[row[1]] = row[2]
ind = open('CWUR2023.csv',encoding='utf-8')
reader = csv.reader(ind)
for row in reader :
    dicc[row[1]] = row[2]
    
for abs in range(len(a)) :
    if a[abs][1] in dicc :
        a[abs].append(dicc[a[abs][1]])
    else :
        a[abs].append("")

writer.writerows(a)